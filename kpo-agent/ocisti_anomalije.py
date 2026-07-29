#!/usr/bin/env python3
"""
Jednokratna skripta - ciscenje starih "anomalija" iz vec postojece KPO knjige.

Sta radi:
  Prolazi kroz sve redove KPO tabele i, za redove koji su ranije bili
  obelezeni kao anomalija (crvena boja, podebljano, komentar "ANOMALIJA: ..."),
  vraca ih na standardan izgled (isti kao svi ostali redovi) i brise komentar.

  Pravi rezervnu kopiju fajla (*.bak-YYYYMMDD-HHMMSS.xlsx) pre nego sto
  isnimi izmene.

Upotreba:
  python ocisti_anomalije.py
  python ocisti_anomalije.py --kpo KPO_2026.xlsx
"""

import argparse
import os
import shutil
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def nadji_ukupno_red(ws, start_row: int = 14) -> int:
    row = start_row
    while row < start_row + 5000:
        val = ws.cell(row=row, column=2).value
        if isinstance(val, str) and val.strip().upper() == "UKUPNO":
            return row
        if ws.cell(row=row, column=1).value is None and val is None:
            return row
        row += 1
    raise RuntimeError("Nije pronadjen UKUPNO red niti prazan prostor u KPO tabeli.")


def main():
    parser = argparse.ArgumentParser(description="Ciscenje starih anomalija iz KPO knjige")
    parser.add_argument("--kpo", default=os.path.join(BASE_DIR, "KPO_prihodi.xlsx"), help="Putanja KPO knjige")
    parser.add_argument("--start-row", type=int, default=14, help="Prvi red sa podacima")
    args = parser.parse_args()

    if not os.path.exists(args.kpo):
        print(f"[GRESKA] Fajl nije pronadjen: {args.kpo}")
        return

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    backup = args.kpo.replace(".xlsx", f".bak-{datetime.now():%Y%m%d-%H%M%S}.xlsx")
    shutil.copy2(args.kpo, backup)
    print(f"Rezervna kopija napravljena: {backup}")

    wb = openpyxl.load_workbook(args.kpo)
    ws = wb.active

    ukupno_row = nadji_ukupno_red(ws, args.start_row)

    fill_ab = PatternFill("solid", fgColor="FFFFCC99")
    fill_e  = PatternFill("solid", fgColor="FFC6EFCE")

    ocisceno = 0
    for row in range(args.start_row, ukupno_row):
        c1 = ws.cell(row=row, column=1)
        c2 = ws.cell(row=row, column=2)

        fgcolor = None
        if c1.fill and c1.fill.fgColor:
            fgcolor = (c1.fill.fgColor.rgb or "").upper()
        ima_komentar_anomalije = bool(c2.comment) and "ANOMALIJA" in (c2.comment.text or "").upper()

        if fgcolor != "FFFF0000" and not ima_komentar_anomalije:
            continue

        for col in (1, 2, 3, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill_ab
            if cell.font and cell.font.bold:
                cell.font = Font(size=cell.font.size or 11, name=cell.font.name or "Calibri", bold=False)

        cell5 = ws.cell(row=row, column=5)
        cell5.fill = fill_e
        if cell5.font and cell5.font.bold:
            cell5.font = Font(size=cell5.font.size or 11, name=cell5.font.name or "Calibri", bold=False)

        if c2.comment:
            c2.comment = None

        ocisceno += 1

    if ocisceno:
        wb.save(args.kpo)
        print(f"Ociscen(o) {ocisceno} red(ova) - vraceni na standardan izgled, komentari uklonjeni.")
        print(f"Sacuvano: {args.kpo}")
    else:
        print("Nije pronadjen nijedan red obelezen kao anomalija - nema sta da se cisti.")


if __name__ == "__main__":
    main()
