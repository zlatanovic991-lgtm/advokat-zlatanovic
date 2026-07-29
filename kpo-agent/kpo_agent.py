#!/usr/bin/env python3
"""
KPO agent - inkrementalno azuriranje Knjige o ostvarenom prometu (KPO)

Sta radi:
  1. Otvara postojecu KPO Excel knjigu i cita poslednji unet datum i redni broj.
  2. Preuzima sa Yahoo Maila samo izvode nakon tog datuma (Raiffeisen + Banca Intesa).
  3. Parsira prilive (odobrenja) iz tih izvoda.
  4. Dodaje nove prilive u KPO knjigu, nastavljajuci numeraciju.
  5. Ako nema novih priliva - javlja da je KPO azurna.

Upotreba:
  python kpo_agent.py
  python kpo_agent.py --od 2026-01-01          (prisilan pocetni datum, npr. prvo pokretanje)
  python kpo_agent.py --kpo KPO_2026.xlsx

Lozinka za Yahoo Mail: environment promenljiva YAHOO_APP_PASS (ili --lozinka, ili unos na tastaturi).
"""

import argparse
import email
import imaplib
import os
import re
import shutil
import sys
import tempfile
from datetime import date, datetime, timedelta
from getpass import getpass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

IMAP_HOST  = "imap.mail.yahoo.com"
IMAP_PORT  = 993
EMAIL_ADDR = "advokatzlatanovic@yahoo.com"

BANKE = {
    "raiffeisen": {
        "sender":  "RaiffeisenOnline@raiffeisenbank.rs",
        "folders": ["INBOX", "izvodi"],
    },
    "intesa": {
        "sender":  "info@mail.bancaintesa.rs",
        "folders": ["INBOX"],
    },
}

KPO_HEADER = {
    "pib":     "111355904",
    "obveznik":"Vladimir Zlatanović",
    "firma":   "Advokat Vladimir Zlatanović",
    "sediste": "Srpskih vladara 99, 18300 Pirot",
    "sifra_delatnosti": "6910",
}


def srpski_broj(x: float) -> str:
    """Formatira broj u srpski format: 45000.0 -> '45.000,00'"""
    s = f"{x:,.2f}"
    return s.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def log_append(log_putanja: str, linija: str) -> None:
    sada = datetime.now().strftime("%Y-%m-%d %H:%M")
    with open(log_putanja, "a", encoding="utf-8") as f:
        f.write(f"{sada} | {linija}\n")


# ─── IMAP ────────────────────────────────────────────────────────────────────

def connect(password: str) -> imaplib.IMAP4_SSL:
    conn = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    conn.login(EMAIL_ADDR, password)
    return conn


def fetch_emails(conn, sender: str, folders: list, od: date, do: date) -> list:
    messages = []
    od_imap = (od - timedelta(days=1)).strftime("%d-%b-%Y")
    do_imap = (do + timedelta(days=2)).strftime("%d-%b-%Y")
    for folder in folders:
        try:
            status, _ = conn.select(f'"{folder}"')
            if status != "OK":
                continue
            criteria = f'(FROM "{sender}" SINCE "{od_imap}" BEFORE "{do_imap}")'
            status, data = conn.search(None, criteria)
            if status != "OK" or not data[0]:
                continue
            for uid in data[0].split():
                status, msg_data = conn.fetch(uid, "(RFC822)")
                if status != "OK":
                    continue
                msg = email.message_from_bytes(msg_data[0][1])
                messages.append(msg)
        except Exception as e:
            print(f"  [upozorenje] {folder}: {e}")
    return messages


def extract_pdf(msg):
    for part in msg.walk():
        ct    = part.get_content_type()
        fname = part.get_filename("") or ""
        if ct == "application/pdf" or fname.lower().endswith(".pdf"):
            return part.get_payload(decode=True)
    return None


# ─── Parsiranje brojeva ──────────────────────────────────────────────────────

def parse_iznos(s):
    """Parsira srpski format broja: 1.234,56 ili 1234.56 ili 1,234.56"""
    if not s:
        return None
    s = str(s).strip().replace("\n", "").replace(" ", "")
    if not s:
        return None
    has_dot = "." in s
    has_comma = "," in s
    if has_dot and has_comma:
        if s.rfind(".") > s.rfind(","):
            s = s.replace(",", "")
        else:
            s = s.replace(".", "").replace(",", ".")
    elif has_comma and not has_dot:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) == 3 and parts[1].isdigit():
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    elif has_dot and not has_comma:
        parts = s.split(".")
        if len(parts) == 2 and len(parts[1]) == 3 and parts[1].isdigit():
            s = s.replace(".", "")
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


# ─── Parsiranje izvoda ───────────────────────────────────────────────────────
# Napomena o formatima: oba pomenuta PDF izvoda (Raiffeisen i Banca Intesa)
# koriste PDF generatore koji ponekad "gube" razmake u ekstraktovanom tekstu
# (npr. "IZVODBR.58" umesto "IZVOD BR. 58"). Regexi ispod su tolerantni na to.

def parse_intesa(pdf_bytes: bytes) -> dict:
    """
    Banca Intesa - klasican format "IZVOD BR." / "PROMENE" tabela (9 kolona):
      RN | Naziv i sediste primaoca-platioca + Broj racuna
         | Poreklo naloga + Datum prijema/knjizenja/valute
         | Zaduzenje | Odobrenje | Sifra | Svrha doznake
         | Poziv na broj (zaduzenje/odobrenje) | Podaci za reklamaciju
    """
    result = {"banka": "Intesa", "datum_izvoda": "", "broj_izvoda": "", "transakcije": []}
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes); tmp_path = tmp.name
    try:
        import pdfplumber
        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                m_br  = re.search(r"IZVOD\s*BR\.?\s*(\d+)", text, re.IGNORECASE)
                m_dat = re.search(r"NA\s*DAN\s*(\d{2}\.\d{2}\.\d{4})", text, re.IGNORECASE)
                if m_br and not result["broj_izvoda"]:
                    result["broj_izvoda"] = m_br.group(1)
                if m_dat and not result["datum_izvoda"]:
                    result["datum_izvoda"] = m_dat.group(1)

                for table in page.extract_tables():
                    for row in table:
                        if not row or len(row) != 9:
                            continue
                        rn = (row[0] or "").strip()
                        if rn in ("", "RN"):
                            continue

                        naziv_cell  = row[1] or ""
                        datumi_cell = row[2] or ""
                        zaduz_cell  = row[3] or ""
                        odob_cell   = row[4] or ""
                        svrha_cell  = row[6] or ""
                        poziv_cell  = row[7] or ""

                        dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", datumi_cell)
                        datum = dates[1] if len(dates) >= 2 else (dates[0] if dates else result["datum_izvoda"])

                        naziv = naziv_cell.split("\n")[0].strip()

                        result["transakcije"].append({
                            "datum":    datum,
                            "naziv":    naziv,
                            "svrha":    svrha_cell.replace("\n", " ").strip(),
                            "pbz_pbo":  poziv_cell.replace("\n", " ").strip(),
                            "na_teret": parse_iznos(zaduz_cell),
                            "u_korist": parse_iznos(odob_cell),
                        })
    finally:
        os.unlink(tmp_path)
    return result


def parse_raiffeisen(pdf_bytes: bytes) -> dict:
    """
    Raiffeisen banka - dvojezicni format "IZVOD / CUSTOMER STATEMENT" (7 kolona):
      RB | Datum prijema/izvrsenja + Broj naloga | Naziv primaoca-posiljaoca + Broj racuna
         | Sifra placanja + Svrha | PBZ/PBO (poziv na broj) | Na teret | U korist
    """
    result = {"banka": "Raiffeisen", "datum_izvoda": "", "broj_izvoda": "", "transakcije": []}
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes); tmp_path = tmp.name
    try:
        import pdfplumber
        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                m_dat = re.search(r"Datum:\s*(\d{2}\.\d{2}\.\d{4})", text)
                m_br  = re.search(r"Broj:\s*(\d+)", text)
                if m_dat and not result["datum_izvoda"]:
                    result["datum_izvoda"] = m_dat.group(1)
                if m_br and not result["broj_izvoda"]:
                    result["broj_izvoda"] = m_br.group(1)

                for table in page.extract_tables():
                    for row in table:
                        if not row or len(row) != 7:
                            continue
                        rn = (row[0] or "").strip()
                        if rn in ("", "RB", "RB:", "No", "No:"):
                            continue

                        datumi_cell = row[1] or ""
                        naziv_cell  = row[2] or ""
                        sifsvr_cell = row[3] or ""
                        poziv_cell  = row[4] or ""
                        teret_cell  = row[5] or ""
                        korist_cell = row[6] or ""

                        dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", datumi_cell)
                        datum = dates[0] if dates else result["datum_izvoda"]

                        naziv = naziv_cell.split("\n")[0].strip()

                        svrha_lines = sifsvr_cell.split("\n")
                        svrha = " ".join(svrha_lines[1:]).strip() if len(svrha_lines) > 1 else sifsvr_cell.strip()

                        result["transakcije"].append({
                            "datum":    datum,
                            "naziv":    naziv,
                            "svrha":    svrha,
                            "pbz_pbo":  poziv_cell.replace("\n", " ").strip(),
                            "na_teret": parse_iznos(teret_cell),
                            "u_korist": parse_iznos(korist_cell),
                        })
    finally:
        os.unlink(tmp_path)
    return result


PARSERI = {"raiffeisen": parse_raiffeisen, "intesa": parse_intesa}


# ─── Prikupljanje priliva ────────────────────────────────────────────────────

def datum_iz_stringa(s: str):
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None


def prikupi_prilive(svi_izvodi: list, posle_datuma) -> list:
    """Vraca listu priliva (odobrenja) iz izvoda, strogo posle datuma posle_datuma (ako je zadat)."""
    prilivi = []
    for izvod in svi_izvodi:
        for t in izvod["transakcije"]:
            if not (t["u_korist"] and t["u_korist"] > 0):
                continue
            datum_str = t["datum"] or izvod["datum_izvoda"]
            datum_dt = datum_iz_stringa(datum_str)
            if posle_datuma and datum_dt and datum_dt <= posle_datuma:
                continue
            naziv = t["naziv"].split(",")[0].strip() if "," in t["naziv"] else t["naziv"]
            prilivi.append({
                "Datum":        datum_str,
                "Banka":        izvod["banka"],
                "Uplatioc":     naziv,
                "Svrha":        (t["svrha"] or "")[:55],
                "Poziv na br.": t["pbz_pbo"],
                "Iznos (RSD)":  t["u_korist"],
                "_datum_dt":    datum_dt,
            })
    prilivi.sort(key=lambda x: x["_datum_dt"] or date.min)
    return prilivi


# ─── KPO Excel ───────────────────────────────────────────────────────────────

def otvori_ili_napravi_kpo(kpo_putanja: str, predlozak_putanja: str):
    import openpyxl
    from openpyxl.styles import Font

    if not os.path.exists(kpo_putanja):
        if not os.path.exists(predlozak_putanja):
            raise FileNotFoundError(f"Predlozak nije nadjen: {predlozak_putanja}")
        shutil.copy2(predlozak_putanja, kpo_putanja)
        wb = openpyxl.load_workbook(kpo_putanja)
        ws = wb.active

        def set_cell(coord, value):
            ws[coord] = value
            ws[coord].font = Font(size=11, name="Calibri")

        set_cell("B2", KPO_HEADER["pib"])
        set_cell("B3", KPO_HEADER["obveznik"])
        set_cell("B5", KPO_HEADER["firma"])
        set_cell("B6", KPO_HEADER["sediste"])
        set_cell("B9", KPO_HEADER["sifra_delatnosti"])
        ws.cell(row=16, column=1, value="Sastavio:").font = Font(size=11)
        ws.cell(row=16, column=5, value="Odgovorno lice:").font = Font(size=11)
        ws.cell(row=17, column=1, value=KPO_HEADER["obveznik"]).font = Font(size=11)
        wb.save(kpo_putanja)
        return wb, ws, True

    wb = openpyxl.load_workbook(kpo_putanja)
    return wb, wb.active, False


def nadji_ukupno_red(ws, start_row: int = 14) -> int:
    row = start_row
    while row < start_row + 5000:
        val = ws.cell(row=row, column=2).value
        if isinstance(val, str) and val.strip().upper() == "UKUPNO":
            return row
        if ws.cell(row=row, column=1).value is None and val is None:
            # nema jos ni UKUPNO reda - bootstrapujemo prazno stanje ovde
            return row
        row += 1
    raise RuntimeError("Nije pronadjen UKUPNO red niti prazan prostor u KPO tabeli.")


def procitaj_poslednje_stanje(ws, start_row: int = 14):
    """Vraca (poslednji_datum, poslednji_redni_broj, ukupno_red)."""
    ukupno_row = nadji_ukupno_red(ws, start_row)
    last_data_row = ukupno_row - 1

    poslednji_redni = 0
    poslednji_datum = None
    for r in range(start_row, last_data_row + 1):
        rb = ws.cell(row=r, column=1).value
        opis = ws.cell(row=r, column=2).value
        if rb is None or opis is None:
            continue
        poslednji_redni = max(poslednji_redni, int(rb))
        m = re.match(r"(\d{2}\.\d{2}\.\d{4})", str(opis).strip())
        if m:
            d = datum_iz_stringa(m.group(1))
            if d and (poslednji_datum is None or d > poslednji_datum):
                poslednji_datum = d

    return poslednji_datum, poslednji_redni, ukupno_row


def dodaj_prilive_u_kpo(ws, novi_prilivi: list, start_row: int = 14):
    """Ubacuje nove redove neposredno pre UKUPNO reda, nastavlja redni broj i azurira ukupno."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not novi_prilivi:
        return

    _, poslednji_redni, ukupno_row = procitaj_poslednje_stanje(ws, start_row)

    if ws.cell(row=ukupno_row, column=1).value is None and ws.cell(row=ukupno_row, column=2).value is None:
        # prazna KPO - bootstrapujemo UKUPNO red pre umetanja
        ws.cell(row=ukupno_row, column=2, value="UKUPNO")
        ws.cell(row=ukupno_row, column=4, value=0)
        ws.cell(row=ukupno_row, column=5, value=0)

    stari_ukupno = ws.cell(row=ukupno_row, column=5).value or 0

    thin = Side(style="thin", color="000000")
    brd_ab = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_ab = PatternFill("solid", fgColor="FFFFCC99")
    fill_e  = PatternFill("solid", fgColor="FFC6EFCE")

    n = len(novi_prilivi)
    ws.insert_rows(ukupno_row, amount=n)

    for i, r in enumerate(novi_prilivi):
        row = ukupno_row + i
        redni = poslednji_redni + 1 + i
        datum_opis = f"{r['Datum']}  {r['Uplatioc']}"
        if r.get("Svrha"):
            datum_opis += f" — {r['Svrha'][:50]}"
        iznos = r["Iznos (RSD)"]

        c = ws.cell(row=row, column=1, value=redni)
        c.font = Font(size=11, name="Calibri"); c.border = brd_ab
        c.fill = fill_ab; c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=row, column=2, value=datum_opis)
        c.font = Font(size=11, name="Calibri"); c.border = brd_ab
        c.fill = fill_ab; c.alignment = Alignment(vertical="center", wrap_text=True)

        c = ws.cell(row=row, column=3, value=None)
        c.border = brd_ab; c.fill = fill_ab

        c = ws.cell(row=row, column=4, value=iznos)
        c.font = Font(size=11, name="Calibri"); c.border = brd_ab
        c.fill = fill_ab; c.number_format = "#,##0.00"
        c.alignment = Alignment(horizontal="right", vertical="center")

        c = ws.cell(row=row, column=5, value=iznos)
        c.font = Font(size=11, name="Calibri"); c.border = brd_ab
        c.fill = fill_e; c.number_format = "#,##0.00"
        c.alignment = Alignment(horizontal="right", vertical="center")

    novi_ukupno_row = ukupno_row + n
    novi_ukupno = stari_ukupno + sum(r["Iznos (RSD)"] for r in novi_prilivi)
    for col in (4, 5):
        c = ws.cell(row=novi_ukupno_row, column=col, value=novi_ukupno)
        c.font = Font(bold=True, size=11, name="Calibri")
        c.number_format = "#,##0.00"
        c.alignment = Alignment(horizontal="right")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="KPO agent - inkrementalno azuriranje KPO knjige")
    parser.add_argument("--od", help="Prisilan pocetni datum (YYYY-MM-DD), koristi se samo ako je KPO prazna")
    parser.add_argument("--do", default=date.today().isoformat(), help="Datum do (default: danas)")
    parser.add_argument("--lozinka", help="Yahoo app lozinka")
    parser.add_argument("--kpo", default=os.path.join(BASE_DIR, "KPO_prihodi.xlsx"), help="Putanja KPO knjige")
    parser.add_argument("--predlozak", default=os.path.join(BASE_DIR, "KPO_predlozak.xlsx"), help="Putanja praznog KPO obrasca")
    parser.add_argument("--log", default=os.path.join(BASE_DIR, "kpo_log.txt"), help="Putanja log fajla")
    args = parser.parse_args()

    do = datetime.strptime(args.do, "%Y-%m-%d").date()

    print("=" * 70)
    print(" KPO agent - inkrementalno azuriranje")
    print("=" * 70)

    wb, ws, kreirana = otvori_ili_napravi_kpo(args.kpo, args.predlozak)
    if kreirana:
        print(f"Nova KPO knjiga kreirana: {args.kpo}")

    poslednji_datum, poslednji_redni, _ = procitaj_poslednje_stanje(ws)

    if poslednji_datum:
        od = poslednji_datum
        print(f"Poslednji unet datum u KPO: {poslednji_datum.strftime('%d.%m.%Y.')}  (redni broj {poslednji_redni})")
    elif args.od:
        od = datetime.strptime(args.od, "%Y-%m-%d").date() - timedelta(days=1)
        print(f"KPO je prazna. Koristim zadati pocetni datum: {args.od}")
    else:
        od = date(do.year, 1, 1) - timedelta(days=1)
        print(f"KPO je prazna. Podrazumevani pocetak: {date(do.year,1,1).strftime('%d.%m.%Y.')} (prosledite --od za drugi datum)")

    print(f"Trazim izvode od {(od + timedelta(days=1)).strftime('%d.%m.%Y.')} do {do.strftime('%d.%m.%Y.')}\n")

    lozinka = args.lozinka or os.environ.get("YAHOO_APP_PASS")
    if not lozinka:
        if sys.stdin.isatty():
            lozinka = getpass("Yahoo app lozinka: ")
        else:
            raise RuntimeError(
                "YAHOO_APP_PASS nije podesena, a skripta se izvrsava bez interaktivne konzole "
                "(npr. iz Task Scheduler-a). Podesi je kao trajnu USER environment promenljivu."
            )

    print("Konektujem se na Yahoo Mail...")
    conn = connect(lozinka)

    svi_izvodi = []
    for naziv, cfg in BANKE.items():
        print(f"Trazim {naziv.capitalize()} izvode...")
        msgs = fetch_emails(conn, cfg["sender"], cfg["folders"], od + timedelta(days=1), do)
        print(f"  Pronadjeno {len(msgs)} mejl(ova).")
        for msg in msgs:
            pdf_bytes = extract_pdf(msg)
            if not pdf_bytes:
                print(f"  [upozorenje] Bez PDF-a: {msg.get('Subject', '')}")
                continue
            izvod = PARSERI[naziv](pdf_bytes)
            svi_izvodi.append(izvod)
            print(f"  Izvod br. {izvod['broj_izvoda']} ({izvod['datum_izvoda']}) - {len(izvod['transakcije'])} stavki")
    conn.logout()
    print()

    # Samo izvodi ciji je datum posle poslednjeg unetog datuma u KPO racunaju se kao "novi"
    svi_izvodi = [
        iz for iz in svi_izvodi
        if not poslednji_datum or not datum_iz_stringa(iz["datum_izvoda"])
        or datum_iz_stringa(iz["datum_izvoda"]) > poslednji_datum
    ]
    svi_izvodi.sort(key=lambda iz: datum_iz_stringa(iz["datum_izvoda"]) or date.min)

    if not svi_izvodi:
        log_append(args.log, "Nema novih izvoda")
        print("Nema novih izvoda - KPO je azurna.")
        return

    ukupno_dodatih_redova = 0

    for izvod in svi_izvodi:
        prilivi = prikupi_prilive([izvod], poslednji_datum)
        iznos_izvoda = sum(p["Iznos (RSD)"] for p in prilivi)
        oznaka_izvoda = f"{izvod['banka']} izvod br.{izvod['broj_izvoda']} ({izvod['datum_izvoda']})"

        if prilivi:
            dodaj_prilive_u_kpo(ws, prilivi)
            wb.save(args.kpo)
            _, poslednji_redni, _ = procitaj_poslednje_stanje(ws)
            ukupno_dodatih_redova += len(prilivi)
            linija = f"{oznaka_izvoda} | +{srpski_broj(iznos_izvoda)} RSD | KPO red {poslednji_redni}"
        else:
            linija = f"{oznaka_izvoda} | bez novih priliva"

        log_append(args.log, linija)
        print(linija)

    print()
    if ukupno_dodatih_redova:
        print(f"KPO azurirana: {args.kpo}  (+{ukupno_dodatih_redova} redova)")
    else:
        print("Novi izvodi pronadjeni, ali bez priliva - KPO nije menjana.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[GRESKA] {e}")
        import traceback
        traceback.print_exc()
        try:
            log_append(os.path.join(BASE_DIR, "kpo_log.txt"), f"GRESKA: {e}")
        except Exception:
            pass
        sys.exit(1)
